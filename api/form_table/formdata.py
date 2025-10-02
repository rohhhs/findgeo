from typing import Dict, List, Optional, Iterable, Any
import os
import pandas as pd
import csv
from openpyxl import load_workbook

from ..definitions import *

def _to_str_row(row: Iterable[Any]) -> List[str]:
    """Convert any row (iterable) to list of strings, normalizing None -> ''"""
    return [("" if v is None else str(v)) for v in row]

def _sniff_csv_dialect(sample: str):
    """Try to sniff delimiter; fallback to common ones."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        # Fallback heuristics: count candidates in sample header line
        # choose the delimiter with most occurrences among common ones
        candidates = [',',';','\t','|']
        scores = {d: sample.count(d) for d in candidates}
        best = max(scores, key=scores.get)
        class SimpleDialect(csv.Dialect):
            delimiter = best
            quotechar = '"'
            doublequote = True
            skipinitialspace = True
            lineterminator = '\n'
            quoting = csv.QUOTE_MINIMAL
        return SimpleDialect()

def readDataBase(
    filepath: str,
    start: Optional[int] = 0,
    amount: Optional[int] = 10000,
    columnKey: Optional[str] = None,
    fieldValue: Optional[str] = None,
    secondColumnKey: Optional[str] = None,
    header_override: Optional[List[str]] = None
) -> List[List[str]]:
    """
    Read .csv or .xlsx and return data as List of Lists.
    - First element is header (List[str]).
    - Subsequent elements are rows (List[str]) corresponding to header columns.
    - `start` and `amount` are applied to data rows (header excluded).
    - If `columnKey` provided and `fieldValue` is None -> return only that column values (header becomes [columnKey]).
    - If `columnKey` + `fieldValue` provided and `secondColumnKey` is None -> return full rows that match (header preserved).
    - If `columnKey` + `fieldValue` + `secondColumnKey` provided -> return only column `secondColumnKey` for rows matching the filter (header becomes [secondColumnKey]).
    - Always returns a list. If any errors/warnings occur, a final element ["__LOG__", "message..."] will be appended.
    """
    logs: List[str] = []
    result: List[List[str]] = []

    if not isinstance(filepath, str) or not filepath:
        logs.append("invalid filepath argument")
        return [["__LOG__","; ".join(logs)]]

    if not os.path.exists(filepath):
        logs.append(f"file not found: {filepath}")
        return [["__LOG__","; ".join(logs)]]

    _, ext = os.path.splitext(filepath.lower())

    # Helper to append final log element if needed
    def _append_log_and_return(rows: List[List[str]]) -> List[List[str]]:
        if logs:
            rows.append(["__LOG__", " | ".join(logs)])
        return rows

    try:
        # ===== XLSX handling =====
        if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            wb = load_workbook(filename=filepath, read_only=True, data_only=True)
            ws = wb.active  # default to active sheet
            rows_iter = ws.iter_rows(values_only=True)
            # skip blank rows at top until we find something to treat as header (unless header_override is provided)
            if header_override:
                header = header_override
            else:
                header = None
                for r in rows_iter:
                    if r is None:
                        continue
                    # check row not completely empty
                    if any(cell is not None and str(cell).strip() != "" for cell in r):
                        header = _to_str_row(r)
                        break
                if header is None:
                    logs.append("xlsx appears empty: no header found")
                    return _append_log_and_return([["__LOG__"," | ".join(logs)]])

            # collect data rows after header
            collected = 0
            skipped = 0
            data_rows: List[List[str]] = []
            for r in rows_iter:
                # treat blank rows as empty row
                row = _to_str_row(r)
                # apply start/amount on data rows
                if skipped < (start or 0):
                    skipped += 1
                    continue
                if amount is not None and collected >= amount:
                    break
                data_rows.append(row)
                collected += 1

            # Now apply columnKey / fieldValue / secondColumnKey logic
            # Normalize header length and row lengths (pad with '' if row shorter)
            max_cols = max(len(header), max((len(r) for r in data_rows), default=0))
            header = header + [""]*(max_cols - len(header))
            normalized_rows = [r + [""]*(max_cols - len(r)) for r in data_rows]

        # ===== CSV handling =====
        elif ext in (".csv", ".txt"):
            # read first chunk to sniff delimiter
            with open(filepath, "r", newline='', encoding="utf-8") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                if header_override:
                    header = header_override
                    # start reading from beginning but skip header detection
                    reader = csv.reader(fh, dialect=_sniff_csv_dialect(sample))
                    # consume the first non-empty row as part of data (because header_override used)
                    rows_iter_list = list(reader)
                    # rows_iter_list may contain header row from file; start/pagination is applied to rows_iter_list from first line
                    data_rows_list = rows_iter_list
                else:
                    try:
                        dialect = _sniff_csv_dialect(sample)
                    except Exception as e:
                        dialect = csv.excel
                        logs.append(f"csv sniff failed, fallback to excel dialect: {e}")
                    reader = csv.reader(fh, dialect=dialect)
                    # find first non-empty row for header
                    header = None
                    rows_buffer = []
                    for r in reader:
                        if header is None:
                            if any((cell is not None and str(cell).strip() != "") for cell in r):
                                header = _to_str_row(r)
                                break
                            else:
                                continue
                    if header is None:
                        logs.append("csv appears empty: no header found")
                        return _append_log_and_return([["__LOG__"," | ".join(logs)]])
                    # After header, the rest are data rows. We need to continue reading from file:
                    # Continue reading remaining rows from fh using same reader
                    data_rows_list = []
                    for r in reader:
                        data_rows_list.append(_to_str_row(r))

            # Now apply start/amount to data_rows_list
            skipped = 0
            collected = 0
            normalized_rows = []
            for r in data_rows_list:
                if skipped < (start or 0):
                    skipped += 1
                    continue
                if amount is not None and collected >= amount:
                    break
                normalized_rows.append(r)
                collected += 1

            # normalize header/rows lengths
            max_cols = max(len(header), max((len(r) for r in normalized_rows), default=0))
            header = header + [""]*(max_cols - len(header))
            normalized_rows = [r + [""]*(max_cols - len(r)) for r in normalized_rows]

        else:
            logs.append(f"unsupported file extension: {ext}")
            return _append_log_and_return([["__LOG__", " | ".join(logs)]])

        # ===== Apply columnKey / fieldValue / secondColumnKey logic on normalized_rows =====
        # If columnKey provided but not in header -> error but still return header and maybe empty data
        if columnKey is not None:
            if columnKey not in header:
                logs.append(f"columnKey '{columnKey}' not found in header")
                # if header_override provided and columnKey missing maybe user expected different header
                # keep going but no rows will match
                col_idx = None
            else:
                col_idx = header.index(columnKey)
        else:
            col_idx = None

        if secondColumnKey is not None:
            if secondColumnKey not in header:
                logs.append(f"secondColumnKey '{secondColumnKey}' not found in header")
                second_idx = None
            else:
                second_idx = header.index(secondColumnKey)
        else:
            second_idx = None

        output_rows: List[List[str]] = []

        # Cases:
        # 1) columnKey is provided and fieldValue is None -> return single-column result of columnKey
        if columnKey is not None and fieldValue is None:
            if col_idx is None:
                # cannot fetch values; return header + no rows
                result_header = [columnKey]
                output_rows = []
            else:
                result_header = [columnKey]
                for r in normalized_rows:
                    output_rows.append([r[col_idx]])
        # 2) columnKey + fieldValue provided
        elif columnKey is not None and fieldValue is not None:
            # filter rows where r[col_idx] == fieldValue (string compare)
            if col_idx is None:
                # no column index -> no matches
                matched = []
            else:
                matched = [r for r in normalized_rows if (r[col_idx] == str(fieldValue))]
            if secondColumnKey is not None:
                # user requested secondColumnKey values from matched rows
                if second_idx is None:
                    # cannot return requested column
                    result_header = [secondColumnKey]
                    output_rows = []
                else:
                    result_header = [secondColumnKey]
                    for r in matched:
                        output_rows.append([r[second_idx]])
            else:
                # return full rows (with header preserved)
                result_header = header
                output_rows = matched
        # 3) no columnKey -> return full rows (header preserved)
        else:
            result_header = header
            output_rows = normalized_rows

        # Final result: [header, ...rows]
        final = [result_header] + output_rows
        return _append_log_and_return(final)

    except Exception as exc:
        # Unexpected error: include traceback-ish message in logs and return it as log-only response
        import traceback
        logs.append("unexpected error: " + repr(exc))
        tb = traceback.format_exc(limit=5)
        logs.append("trace: " + tb.replace("\n", " | "))
        return [["__LOG__", " | ".join(logs)]]

def formDataBase(filepath: str, headkeys: Dict[str, list], newfilename: str) -> str:
    """
    Reform a database by standardizing headers using synonyms provided in a dictionary.
    
    Args:
        filepath: Path to the input database file (.csv or .xlsx)
        headkeys: Dictionary mapping standard headers to lists of possible synonyms
        newfilename: Name for the output .csv file
    
    Returns:
        Path to the new .csv file with standardized headers
    """
    # Determine file extension and read the file
    _, ext = os.path.splitext(filepath)
    if ext.lower() == '.xlsx':
        df = pd.read_excel(filepath)
    elif ext.lower() == '.csv':
        df = pd.read_csv(filepath)
    else:
        raise ValueError(f"Unsupported file format: {ext}")
    
    # Create a mapping from synonyms to standard headers
    synonym_to_header = {}
    for standard_header, synonyms in headkeys.items():
        for synonym in synonyms:
            synonym_to_header[synonym.lower()] = standard_header
    
    # Identify columns that match synonyms
    matched_columns = {}
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in synonym_to_header:
            matched_columns[col] = synonym_to_header[col_lower]
    
    # Create a new DataFrame with standardized headers
    new_df = pd.DataFrame()
    
    for standard_header in headkeys.keys():
        # Find if any of the original columns match this standard header
        matched_col = None
        for original_col, mapped_header in matched_columns.items():
            if mapped_header == standard_header:
                matched_col = original_col
                break
        
        if matched_col is not None:
            new_df[standard_header] = df[matched_col]
        else:
            # If no matching column found, create an empty column
            new_df[standard_header] = pd.NA
    
    # Write the new DataFrame to a CSV file
    new_filepath = newfilename if newfilename.endswith('.csv') else newfilename + '.csv'
    new_df.to_csv(new_filepath, index=False)
    
    return new_filepath

if __name__ == "__main__":
    print ("working")
    # Example 1: read first 100 data rows from CSV/XLSX
    # rows = readDataBase(os.path.join(ASSET_DATA_DIR , 'test.csv'), start=0, amount=100)
    # print (rows)

    # with open(f'./api/form_table/jsondata002.json', 'w', encoding="utf-8") as f:
    #     f.write(str(rows))

    # rows[0] is the header list; rows[1:] are the data rows.

    # # Example 2: return only the column 'filename' values starting from the 10th data row
    # filenames = readDataBase(os.path.join(ASSET_DATA_DIR , 'test.csv'), start=10, amount=50, columnKey="filename")
    # print (filenames)

    # with open(f'./api/form_table/jsondata002-01.json', 'w', encoding="utf-8") as f:
    #     f.write(str(filenames))

    # # Example 3: find rows where 'camera' == "2f8f5351-f452-4214-954d-ea0bebfa02e4" and return entire rows
    # matches = readDataBase(os.path.join(ASSET_DATA_DIR , 'test.csv'), columnKey="camera", fieldValue="2f8f5351-f452-4214-954d-ea0bebfa02e4")
    # print (matches)

    # with open(f'./api/form_table/jsondata002-02.json', 'w', encoding="utf-8") as f:
    #     f.write(str(matches))

    # # Example 4: find row(s) where 'camera' == "2f8f5351-f452-4214-954d-ea0bebfa02e4" and return the value of 'filename' column for those rows
    # vals = readDataBase(os.path.join(ASSET_DATA_DIR , 'test.csv'), columnKey="camera", fieldValue="2f8f5351-f452-4214-954d-ea0bebfa02e4", secondColumnKey="filename")
    # print (vals)

    # with open(f'./api/form_table/jsondata002-03.json', 'w', encoding="utf-8") as f:
    #     f.write(str(vals))

    # # Example 5: headless CSV: you know header externally . Useful to override column name formating with same row field length .
    # rows = readDataBase(os.path.join(ASSET_DATA_DIR , 'test.csv'), start=1, amount=50, header_override=["filename","camera","latitude","batat"])
    # print (rows)

    # with open(f'./api/form_table/jsondata002-04.json', 'w', encoding="utf-8") as f:
    #     f.write(str(rows))